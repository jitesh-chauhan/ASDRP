import openpyxl.workbook
import pandas as pd
from django.shortcuts import render
import os
import imaplib
import email
from .models import RawData, ProcessedData
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.conf import settings
import smtplib
from email.message import EmailMessage

# Create your views here.


def fetchEmail(request):
    RawData.objects.all().delete()

    def emailData():
        imap_server = "outlook.office365.com"
        username = ""  # your email
        password = ""
        port = 993

        mail = imaplib.IMAP4_SSL(imap_server, port)
        mail.login(username, password)
        print("Login successful!")

        mail.select("inbox")
        status, messages = mail.search(None, f'SUBJECT "Daily Sales Report"')
        emails = []

        if status == "OK":
            email_ids = messages[0].split()
            for e_id in email_ids:
                status, msg_data = mail.fetch(e_id, "(RFC822)")
                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        subject = msg["subject"]
                        print(f"Email Subject: {subject}")
                        emails.append(msg)
        mail.close()
        mail.logout()
        return emails, username, subject

    def download_attachments(emails, download_folder, subject):
        if not os.path.exists(download_folder):
            os.makedirs(download_folder)
        print(subject)
        downloaded_filename = None

        for msg in emails:
            if subject == "Daily Sales Report":
                for part in msg.walk():
                    if (
                        part.get_content_maintype() == "multipart"
                        or part.get("Content-Disposition") is None
                    ):
                        continue
                    filename = part.get_filename()
                    if filename and (
                        filename.endswith(".xlsx") or filename.endswith(".xls")
                    ):
                        filepath = os.path.join(download_folder, filename)
                        with open(filepath, "wb") as f:
                            f.write(part.get_payload(decode=True))
                        print(f"Downloaded {filename} to {download_folder}")
                        downloaded_filename = filename
                        return downloaded_filename

        return downloaded_filename

    (
        emails,
        username,
        subject,
    ) = emailData()

    if emails:
        filename = download_attachments(emails, r"Files", subject)
    else:
        print("No emails found with the given keyword.")

    def read_and_store_latest_xls(filename):

        df = pd.read_excel(f"Files/{filename}")
        if len(df) <= 8:
            raise ValueError("Not enough data")
        unique = RawData.objects.filter()

        for index, row in df.iterrows():

            existing_data = RawData.objects.filter(
                Store_ID=row["Store_ID"],
                Date=row["Date"],
                Product_ID=row["Product_ID"],
                Quantity_Sold=row["Quantity_Sold"],
                Total_Sales=row["Total_Sales"],
            )

            if not existing_data.exists():
                sales_data = RawData.objects.create(
                    Store_ID=row["Store_ID"],
                    Date=row["Date"],
                    Product_ID=row["Product_ID"],
                    Quantity_Sold=row["Quantity_Sold"],
                    Total_Sales=row["Total_Sales"],
                )
                sales_data.save()

    try:
        read_and_store_latest_xls(filename)
    except (FileNotFoundError, ValueError) as e:
        print(e)

    def CalculateData():

        unique_combinations = RawData.objects.values("Store_ID", "Date").distinct()

        for combination in unique_combinations:
            store_id = combination["Store_ID"]
            date = combination["Date"]

            sales_data = RawData.objects.filter(Store_ID=store_id, Date=date)

            total_sales = 0
            top_selling_product = None
            max_quantity_sold = -1

            for data in sales_data:
                total_sales += data.Total_Sales
                if data.Quantity_Sold > max_quantity_sold:
                    max_quantity_sold = data.Quantity_Sold
                    top_selling_product = data.Product_ID

            existing_data = ProcessedData.objects.filter(
                Store_ID=store_id,
                Date=date,
                Total_Sales=total_sales,
                Top_Selling_Product=top_selling_product,
            )
            if not existing_data.exists():
                ProcessedData.objects.update_or_create(
                    Store_ID=store_id,
                    Date=date,
                    Total_Sales=total_sales,
                    Top_Selling_Product=top_selling_product,
                )

                print("data processed")
            
            

    CalculateData()

    data = ProcessedData.objects.all().values()

    def ExportData():
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "ProcessedData"

        fields = ["Store_ID", "Date", "Total_Sales", "Top_Selling_Product"]

        for col_num, field in enumerate(fields, 1):
            column_letter = get_column_letter(col_num)
            worksheet[f"{column_letter}1"] = field

        for row_num, processed_data in enumerate(ProcessedData.objects.all(), 2):
            worksheet[f"A{row_num}"] = processed_data.Store_ID
            worksheet[f"B{row_num}"] = processed_data.Date
            worksheet[f"C{row_num}"] = processed_data.Total_Sales
            worksheet[f"D{row_num}"] = processed_data.Top_Selling_Product

            date = processed_data.Date

        file_dir = os.path.join(settings.BASE_DIR, "Files/ProcessedData")
        os.makedirs(file_dir, exist_ok=True)
        file_path = os.path.join(file_dir, f"ProcessedFile{date}.xlsx")
        if not os.path.exists(file_path):
            workbook.save(file_path)
        else:
            print("file already exists")
        return file_path

    file_path = ExportData()

    def SendFile(username):
        # username = 'jiteshchauhan1234@outlook.com'
        receiver_email = [
            "example@gmail.com",
            "receiver@example.com",
        ]  # email id of the receiver
        subject = "Processed Data summary"
        body = "Please find the attached file"

        msg = EmailMessage()
        msg["From"] = username
        msg["To"] = ",".join(receiver_email)
        msg["Subject"] = subject
        msg.set_content(body)

        with open(file_path, "rb") as f:
            file_data = f.read()
            msg.add_attachment(
                file_data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(file_path),
            )

        try:
            with smtplib.SMTP(
                "smtp-mail.outlook.com", 587
            ) as smtp:  # replace with your smtp-server address and port
                smtp.starttls()
                smtp.login(username, "")  # replace with your password
                smtp.send_message(msg)
            print("Email sent successfully!")
        except Exception as e:
            print(f"Error sending email: {e}")

    SendFile(username)

    return render(request, "FetchEmail.html", {"data": data})


def Home(request):

    return render(request, "home.html")
